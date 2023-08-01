import inspect
import logging
import openpyxl


class Utils:
    def assertlistitemtext(self,list, value):
        for stop in list:
            print('The text is:', stop.text)
            assert stop.text == value
            print("assert pass")

    def custom_logger(logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]

        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler("automation.log")
        formatter = logging.Formatter("%(asctime)s - %(levelname)s: %(message)s", datefmt="%d/%m/%Y %I:%M:%S %p")
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(path, sheet):
        datalist = []
        wb = openpyxl.load_workbook(path)
        sh = wb[sheet]
        rc = sh.max_row
        cc = sh.max_column
        for i in range(2, rc + 1):
            row = []
            for j in range(1, cc + 1):
                row.append(sh.cell(i, j).value)
            datalist.append(row)
        return datalist
